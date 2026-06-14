"""CRM import wizard core — parse → map → dedup → commit.

The wizard accepts an uploaded ``.csv``/``.xlsx`` file *or* a Google Sheet URL,
maps the source headers onto CRM fields, splits rows into new vs already-known
(dedup), and commits the new rows into Organization/Contact/OutreachThread.

Design notes
------------
- ``openpyxl`` is imported **lazily inside** ``parse_rows`` (mirrors the lazy
  Google-Sheets imports in ``apps/content_intake/sheets_sync.py``) so Celery's
  autodiscover never has to import the heavy library. The prod Docker image
  installs it from ``requirements.txt``.
- ``parse_sheet_url`` reuses the content_intake grid reader
  (``_get_sheet_grid``) — re-exported here so tests can patch
  ``apps.crm.import_wizard._get_sheet_grid``.
- ``commit_rows`` wraps every row in its own try/except → a per-row result
  ``{row, status, error?}``. A bad row (e.g. missing org_name) becomes an
  ``error`` result and is NEVER a silent drop or a crash.

The CRM target fields the wizard understands:
    org_name, contact_name, contact_email, role, stage, track
"""
from __future__ import annotations

import csv
import io
import re

# Re-exported so the grid reader can be patched at this module's namespace.
from apps.content_intake.sheets_sync import _get_sheet_grid  # noqa: F401

# Fields a mapping value may target. ``org_name`` is required at commit time.
CRM_FIELDS = ("org_name", "contact_name", "contact_email", "role", "stage", "track")

_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_rows(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse an uploaded ``.csv`` / ``.xlsx`` into a list of header→value dicts.

    The first row is treated as the header. Raises ``ValueError`` for an
    unsupported extension.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(file_bytes)
    if name.endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    raise ValueError(
        f"Unsupported file type: {filename!r}. Upload a .csv or .xlsx, or paste a Google Sheet URL."
    )


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")  # tolerate a BOM from Excel exports
    reader = csv.reader(io.StringIO(text))
    grid = [row for row in reader]
    return _rows_from_grid(grid)


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    # Lazy import — keep openpyxl out of Celery autodiscover (requirements gotcha).
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    grid: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append(["" if c is None else str(c) for c in row])
    try:
        wb.close()
    except Exception:
        pass
    return _rows_from_grid(grid)


def _rows_from_grid(grid: list[list[str]]) -> list[dict]:
    """Turn a 2-D grid (row 0 = header) into a list of header→value dicts.

    Trailing fully-blank rows are dropped.
    """
    if not grid:
        return []
    headers = [str(h).strip() for h in grid[0]]
    out: list[dict] = []
    for row in grid[1:]:
        values = [str(c).strip() for c in row]
        if not any(values):
            continue  # skip blank trailing rows
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = values[idx] if idx < len(values) else ""
        out.append(record)
    return out


def parse_sheet_url(url: str, sheet_range: str = "A:Z") -> list[dict]:
    """Parse a Google Sheet URL into header→value dicts via the grid reader.

    Reuses ``apps/content_intake/sheets_sync._get_sheet_grid`` (re-exported in
    this module, so tests patch ``apps.crm.import_wizard._get_sheet_grid``).
    Returns ``[]`` when the sheet id can't be extracted or the sheet is empty.
    """
    sheet_id = _extract_sheet_id(url)
    if not sheet_id:
        return []
    grid_cells = _get_sheet_grid(sheet_id, sheet_range)
    # Each cell is a dict (formattedValue/hyperlink/…); reduce to visible text.
    grid = [[str(c.get("formattedValue", "")).strip() for c in row] for row in grid_cells]
    return _rows_from_grid(grid)


def _extract_sheet_id(url: str) -> str:
    if not url:
        return ""
    m = _SHEET_ID_RE.search(url)
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# Mapping
# ---------------------------------------------------------------------------


def apply_mapping(rows: list[dict], mapping: dict) -> list[dict]:
    """Map source-header dicts onto CRM-field dicts using ``mapping``.

    ``mapping`` is ``{source_header: crm_field}``. Unmapped source columns are
    dropped; mapping targets that aren't CRM fields are ignored.
    """
    mapped: list[dict] = []
    for row in rows:
        out: dict = {}
        for src_header, crm_field in mapping.items():
            if crm_field not in CRM_FIELDS:
                continue
            out[crm_field] = str(row.get(src_header, "") or "").strip()
        mapped.append(out)
    return mapped


# ---------------------------------------------------------------------------
# Dedup
# ---------------------------------------------------------------------------


def dedupe(mapped_rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Split mapped rows into ``(new, matched)``.

    A row is *matched* when an Organization with the same name (case
    -insensitive) already has a Contact with the same email (case-insensitive)
    in the DB. Everything else is *new*. Org-name-only collisions are still
    *new* (a new contact under an existing org).
    """
    from apps.crm.models import Contact

    new: list[dict] = []
    matched: list[dict] = []
    for row in mapped_rows:
        org_name = (row.get("org_name") or "").strip()
        email = (row.get("contact_email") or "").strip()
        is_match = False
        if org_name and email:
            is_match = Contact.objects.filter(
                org__name__iexact=org_name, email__iexact=email
            ).exists()
        (matched if is_match else new).append(row)
    return new, matched


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------


def commit_rows(new_rows: list[dict]) -> list[dict]:
    """Create Organization/Contact/OutreachThread for each mapped row.

    Returns a per-row result list ``[{row, status, error?}]`` where status is
    ``created`` or ``error``. Every row is wrapped in its own transaction so a
    bad row neither crashes the import nor leaves a partial write — it is
    reported, never silently dropped.
    """
    from django.db import transaction

    from apps.crm.models import Contact, Organization, OutreachThread

    results: list[dict] = []
    for row in new_rows:
        org_name = (row.get("org_name") or "").strip()
        if not org_name:
            results.append({"row": row, "status": "error", "error": "missing org_name"})
            continue
        try:
            with transaction.atomic():
                org = Organization.objects.filter(name__iexact=org_name).first()
                if org is None:
                    org = Organization.objects.create(name=org_name)

                contact = None
                email = (row.get("contact_email") or "").strip()
                full_name = (row.get("contact_name") or "").strip()
                if email:
                    contact = Contact.objects.filter(org=org, email__iexact=email).first()
                if contact is None and full_name:
                    contact = Contact.objects.filter(org=org, full_name__iexact=full_name).first()
                if contact is None and (email or full_name):
                    contact = Contact.objects.create(
                        org=org,
                        full_name=full_name,
                        email=email,
                        role=(row.get("role") or "").strip(),
                    )

                OutreachThread.objects.create(
                    org=org,
                    primary_contact=contact,
                    stage=(row.get("stage") or OutreachThread.Stage.TARGETED).strip()
                    or OutreachThread.Stage.TARGETED,
                    track=(row.get("track") or "").strip(),
                )
            results.append({"row": row, "status": "created"})
        except Exception as exc:  # never silently drop a row
            results.append({"row": row, "status": "error", "error": str(exc)})
    return results
